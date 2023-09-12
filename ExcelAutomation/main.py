import openpyxl


def total_column_values(input_file):
    wb = openpyxl.load_workbook(input_file)

    # Get the active sheet (Sheet1 in this case)
    sheet1 = wb.active

    # Get the full as a reference sheet (Sheet2 in this case)
    sheet2 = wb['Sheet2']

    # Calculate the total of values in 'B' column of Sheet1
    total = sum(row[1] for row in sheet1.iter_rows(min_row=2, values_only=True) if row[1])
    total1 = sum(row[6] for row in sheet1.iter_rows(min_row=2, values_only=True) if row[6])
    '''total = 0
    for row in sheet1.iter_rows(min_row=2, values_only=True):  # Header in row 1
        value = row[1]  # Index 1 corresponds to column 'B' (0-based index)
        if value:
            total += value

    total1 = 0
    for row in sheet1.iter_rows(min_row=2, values_only=True):  # Header in row 1
        # Index 1 corresponds to column 'G' (0-based index)
        value1 = row[6]
        if value1:
            total1 += value1'''

    v1 = sheet2.cell(row=2, column=1).value
    if v1 == 'Total':
        sheet2.cell(row=2, column=2, value=total)
    else:
        sheet2.append(['Total', total])

    v2 = sheet2.cell(row=2, column=7).value
    if v2 == 'Total':
        sheet2.cell(row=2, column=8, value=total1)
    else:
        sheet2.append(['Total', total1])

    wb.save(input_file)


if __name__ == "__main__":
    input_file = "text.xlsx"
    total_column_values(input_file)
    print("Total calculated and saved to output.xlsx.")
